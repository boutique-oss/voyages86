#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

BASE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(BASE, 'data', 'package', 'data')

communes = json.load(open(os.path.join(DATA,'communes.json'), encoding='utf-8'))
deps = {d['code']: d for d in json.load(open(os.path.join(DATA,'departements.json'), encoding='utf-8'))}
regs = {r['code']: r for r in json.load(open(os.path.join(DATA,'regions.json'), encoding='utf-8'))}

actuelles = [c for c in communes if c.get('type')=='commune-actuelle']
actuelles.sort(key=lambda c:(c.get('departement',''), c.get('nom','')))

# ---------- styles ----------
ARIAL = 'Arial'
hdr_font = Font(name=ARIAL, bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill('solid', fgColor='1F4E5F')
title_font = Font(name=ARIAL, bold=True, size=16, color='1F4E5F')
sub_font = Font(name=ARIAL, italic=True, size=10, color='555555')
base_font = Font(name=ARIAL, size=10)
input_fill = PatternFill('solid', fgColor='FFF7E0')  # cells to fill in
thin = Side(style='thin', color='D0D0D0')
border = Border(left=thin,right=thin,top=thin,bottom=thin)

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1 : Référentiel (master national, figé)
# ============================================================
ws = wb.active
ws.title = 'Référentiel'
ws['A1'] = 'Référentiel des communes de France'
ws['A1'].font = title_font
ws['A2'] = "Source : Etalab / INSEE — Code Officiel Géographique, millésime 2026 (@etalab/decoupage-administratif v6.0.0). Clé unique : code INSEE. Feuille figée — ne pas modifier."
ws['A2'].font = sub_font
headers = ['code_INSEE','nom','departement','nom_departement','region','nom_region','population']
hrow = 4
for i,h in enumerate(headers,1):
    ws.cell(row=hrow, column=i, value=h)
style_header(ws, hrow, len(headers))
r = hrow+1
for c in actuelles:
    d = deps.get(c.get('departement'),{})
    reg = regs.get(c.get('region'),{})
    ws.cell(row=r,column=1,value=c['code'])
    ws.cell(row=r,column=2,value=c['nom'])
    ws.cell(row=r,column=3,value=c.get('departement'))
    ws.cell(row=r,column=4,value=d.get('nom'))
    ws.cell(row=r,column=5,value=c.get('region'))
    ws.cell(row=r,column=6,value=reg.get('nom'))
    ws.cell(row=r,column=7,value=c.get('population'))
    r += 1
last_ref = r-1
for row in ws.iter_rows(min_row=hrow+1, max_row=last_ref, max_col=7):
    for cell in row:
        cell.font = base_font
widths=[12,34,12,20,9,22,12]
for i,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A5'
ws.auto_filter.ref = f'A{hrow}:G{last_ref}'
print('Référentiel rows:', last_ref-hrow, 'last row', last_ref)

REF_RANGE_CODE = f"Référentiel!$A${hrow+1}:$A${last_ref}"
REF_RANGE_NOM  = f"Référentiel!$B${hrow+1}:$B${last_ref}"

# ============================================================
# Sheet 2 : Suivi_Vienne (dept 86)
# ============================================================
d86 = [c for c in actuelles if c.get('departement')=='86']
ws2 = wb.create_sheet('Suivi_Vienne')
ws2['A1']='Suivi des visites — Vienne (86)'
ws2['A1'].font=title_font
ws2['A2']="Zone de travail. Saisir uniquement les colonnes Statut, Date et Notes (fond crème). Le nom se remplit tout seul depuis le Référentiel via le code INSEE."
ws2['A2'].font=sub_font
h=['code_INSEE','nom (auto)','statut','date_visite','notes']
hr2=4
for i,x in enumerate(h,1): ws2.cell(row=hr2,column=i,value=x)
style_header(ws2,hr2,len(h))

# demo statuses to match the map
demo = {'86194':'Visité','86041':'Visité','86214':'Visité','86027':'Prévu','86133':'Prévu'}
demo_dates = {'86194':'2026-05-10','86041':'2026-05-10','86214':'2026-06-02'}
demo_notes = {'86194':'Centre historique, Notre-Dame-la-Grande','86214':'Abbaye de Saint-Benoît'}

statuts_order = {'Visité':0,'Prévu':1,'À visiter':2}
d86_sorted = sorted(d86, key=lambda c: c['nom'])
r=hr2+1
for c in d86_sorted:
    code=c['code']
    ws2.cell(row=r,column=1,value=code)
    ws2.cell(row=r,column=2,value=f'=IFERROR(INDEX({REF_RANGE_NOM},MATCH(A{r},{REF_RANGE_CODE},0)),"code inconnu")')
    ws2.cell(row=r,column=3,value=demo.get(code,'À visiter'))
    ws2.cell(row=r,column=4,value=demo_dates.get(code,''))
    ws2.cell(row=r,column=5,value=demo_notes.get(code,''))
    r+=1
last86=r-1
# styling + input fills
for row in ws2.iter_rows(min_row=hr2+1,max_row=last86,max_col=5):
    for cell in row: cell.font=base_font; cell.border=border
    row[2].fill=input_fill; row[3].fill=input_fill; row[4].fill=input_fill  # statut,date,notes
for i,w in enumerate([12,26,12,14,40],1):
    ws2.column_dimensions[get_column_letter(i)].width=w
ws2.freeze_panes='A5'
# data validation dropdown for statut
dv=DataValidation(type='list',formula1='"À visiter,Prévu,Visité"',allow_blank=False)
ws2.add_data_validation(dv)
dv.add(f'C{hr2+1}:C{last86}')
print('Suivi_Vienne communes:', last86-hr2)

# ============================================================
# Sheet 3 : Tableau de bord
# ============================================================
ws3=wb.create_sheet('Tableau de bord')
ws3['A1']='Tableau de bord — progression'
ws3['A1'].font=title_font
ws3['A2']="Se met à jour automatiquement quand vous changez les statuts dans la feuille Suivi_Vienne."
ws3['A2'].font=sub_font
ws3['A4']='Département'; ws3['B4']='Vienne (86)'
statut_rng=f"Suivi_Vienne!$C${hr2+1}:$C${last86}"
rows=[
 ('Communes au total', f'=COUNTA(Suivi_Vienne!$A${hr2+1}:$A${last86})'),
 ('Visité',            f'=COUNTIF({statut_rng},"Visité")'),
 ('Prévu',             f'=COUNTIF({statut_rng},"Prévu")'),
 ('À visiter',         f'=COUNTIF({statut_rng},"À visiter")'),
 ('% visité',          '=IFERROR(B6/B5,0)'),
]
rr=5
for label,formula in rows:
    ws3.cell(row=rr,column=1,value=label).font=Font(name=ARIAL,bold=True,size=11)
    ws3.cell(row=rr,column=2,value=formula).font=base_font
    rr+=1
ws3['B9'].number_format='0.0%'
for r_ in range(4,10):
    ws3.cell(row=r_,column=1).border=border; ws3.cell(row=r_,column=2).border=border
ws3.column_dimensions['A'].width=22; ws3.column_dimensions['B'].width=16

# small bar chart of the 3 statuses
chart=BarChart(); chart.type='col'; chart.title='Répartition des statuts — Vienne'
chart.y_axis.title='Communes'; chart.legend=None
data=Reference(ws3,min_col=2,min_row=6,max_row=8)
cats=Reference(ws3,min_col=1,min_row=6,max_row=8)
chart.add_data(data,titles_from_data=False); chart.set_categories(cats)
chart.width=12; chart.height=7
ws3.add_chart(chart,'D4')

# move dashboard first for visibility
wb.move_sheet('Tableau de bord', -(len(wb.sheetnames)-1))

out=os.path.join(BASE,'Carnet_de_voyage_communes.xlsx')
wb.save(out)
print('saved', out)

# also emit suivi.json for the map (same statuses)
suivi={c['code']: demo.get(c['code'],'À visiter') for c in d86_sorted}
json.dump(suivi, open(os.path.join(BASE,'data','suivi-86.json'),'w',encoding='utf-8'), ensure_ascii=False)
print('suivi-86.json written, visited:', sum(1 for v in suivi.values() if v=='Visité'))
